
# -*- coding: utf-8 -*-
import PySimpleGUI as sg
import pandas as pd
import openpyxl
import os


#### Άνοιγμα αρχείων Excel
pedio1 = pd.read_excel (os.getcwd()+'/Αρχεία Εφαρμογής/ΣΥΝΤΕΛΕΣΤΕΣ1.xlsx')
pedio2 = pd.read_excel (os.getcwd()+'/Αρχεία Εφαρμογής/ΣΥΝΤΕΛΕΣΤΕΣ2.xlsx')
pedio3 = pd.read_excel (os.getcwd()+'/Αρχεία Εφαρμογής/ΣΥΝΤΕΛΕΣΤΕΣ3.xlsx')
pedio4 = pd.read_excel (os.getcwd()+'/Αρχεία Εφαρμογής/ΣΥΝΤΕΛΕΣΤΕΣ4.xlsx')

#### Κατευθύνσεις ΓΕΛ
kateythinseis_gel= {"Ανθρωπιστικών Σπουδών - 1ο Πεδίο":
                    [{"Μαθήματα" : "Νεοελληνική Γλώσσα Και Λογοτεχνία"},
                     {"Μαθήματα" : "Αρχαιά Ελληνικά"},
                     {"Μαθήματα" : "Ιστορία"},
                     {"Μαθήματα" : "Λατινικά"}],
                    "Θετικές και Τεχνολογικές Επιστήμες - 2ο Πεδίο":
                     [{"Μαθήματα" : "Νεοελληνική Γλώσσα Και Λογοτεχνία"},
                     {"Μαθήματα" : "Φυσική"},
                     {"Μαθήματα" : "Χημεία"},
                     {"Μαθήματα" : "Μαθηματικά"}],
                    "Επιστήμες Υγείας και Ζωής - 3ο Πεδίο":
                     [{"Μαθήματα" : "Νεοελληνική Γλώσσα Και Λογοτεχνία"},
                     {"Μαθήματα" : "Φυσική"},
                     {"Μαθήματα" : "Χημεία"},
                     {"Μαθήματα" : "Βιολογία"}],
                    "Επιστήμες Οικονομίας και Πληροφορικής - 4ο Πεδίο":
                     [{"Μαθήματα" : "Νεοελληνική Γλώσσα Και Λογοτεχνία"},
                     {"Μαθήματα" : "Μαθηματικά"},
                    {"Μαθήματα" : "Πληροφορική"},
                     {"Μαθήματα" : "Οικονομία"}]}
width_kat_gel = max(map(len, kateythinseis_gel))+1

#### Γραφικό Περιβάλλον
layout = [
    [sg.Text("Μόρια Εισαγωγής Πανελληνίων",font=('Segoe Print',23,"bold"),background_color='White',text_color='#020261')],
    [sg.Image(source=r"Αρχεία Εφαρμογής/ena-photo.png",size=(650,250))],
    [sg.Text("Ονοματεπώνυμο Μαθητή",background_color='White',text_color='Black'),sg.InputText(key="Ονοματεπώνυμο Μαθητή",expand_x=True)],
    [sg.Text("Κατεύθυνση:",background_color='White',text_color='Black'),sg.Combo(list(kateythinseis_gel.keys()), size=(width_kat_gel, 4),expand_x=True,key="Επιλογή Κατεύθυνσης", enable_events=True),sg.Button("OK",key='OK')],
    [sg.Text("Βαθμοί :",font=('Segoe Print',17,"bold"),background_color='White',text_color='#020261')],
    [sg.T("",key='Μάθημα 1',background_color='White',text_color='Black'),sg.InputText(key="Βαθμός1",expand_x=True)],
    [sg.T("",key='Μάθημα 2',background_color='White',text_color='Black'),sg.InputText(key="Βαθμός2",expand_x=True)],
    [sg.T("",key='Μάθημα 3',background_color='White',text_color='Black'),sg.InputText(key="Βαθμός3",expand_x=True)],
    [sg.T("",key='Μάθημα 4',background_color='White',text_color='Black'),sg.InputText(key="Βαθμός4",expand_x=True)],
    [sg.Button("Υπολογισμός",expand_x=True),sg.Button("Καθαρισμός",key='Clear',expand_x=True,enable_events=True),sg.Button("Έξοδος",expand_x=True)],
    [sg.Image(source=r"Αρχεία Εφαρμογής/MV.png",subsample=2)]
]

window = sg.Window("Υπολογιστής Μορίων",layout,resizable=True,font="TimesNewRoman",element_justification='center',background_color= 'white',icon=r"Αρχεία Εφαρμογής/ena-logo.ico",finalize= True)

math=['Μάθημα 1','Μάθημα 2','Μάθημα 3','Μάθημα 4']
grade=["Βαθμός1","Βαθμός2","Βαθμός3","Βαθμός4"]

while True:
    event, values = window.read()
    
##### Έξοδος
    if event == "Έξοδος" or event ==sg.WIN_CLOSED:
        break
    
##### Επιλογή Κατεύθυνσης ΟΚ
    if event == "OK":
        epilogi = values["Επιλογή Κατεύθυνσης"]
        title_list = [i["Μαθήματα"] for i in kateythinseis_gel[epilogi]]
        for i in range(len(title_list)):
           window[math[i]].Update(title_list[i])

##### Καθαρισμός
    if event == "Clear":
        window['Ονοματεπώνυμο Μαθητή'].Update('')
        window['Επιλογή Κατεύθυνσης'].Update('')
        window['Μάθημα 1'].Update('')
        window['Μάθημα 2'].Update('')
        window['Μάθημα 3'].Update('')
        window['Μάθημα 4'].Update('')
        window['Βαθμός1'].Update('')
        window['Βαθμός2'].Update('')
        window['Βαθμός3'].Update('')
        window['Βαθμός4'].Update('')

##### Υπολογισμός
    if event == "Υπολογισμός":
       codename = str(values["Ονοματεπώνυμο Μαθητή"]+".xlsx")
       wrong_values=False
       X=[]

######## Έλεγχος για λάθος τιμές
       
########## Κενό Ονοματεπώνυμο
       if values["Ονοματεπώνυμο Μαθητή"]=='':
           sg.popup("Κενό Ονοματεπώνυμο Μαθητή. Πληκτρολόγησε το Ονοματεπώνυμο του Μαθητή.")
           wrong_values=True
       if wrong_values:
           continue

########## Κενός Βαθμός
       for i in range(4):
           if values[grade[i]] == '':
               sg.popup("Κενός βαθμός μαθήματος. Πληκτρολόγησε έναν αριθμό μεταξύ 0-20.")
               wrong_values=True
               break
       if wrong_values:
           continue

########## Λάθος Βαθμός
       for i in range(4):
           X.insert(i,(float(values[grade[i]])))
           if X[i]<0 or X[i]>20:
               sg.popup("Λάθος βαθμός μαθήματος. Ο αριθμός θα πρέπει να είναι μεταξύ 0-20.")
               wrong_values=True
               break
       if wrong_values:
           continue

######## Ανθρωπιστικών Σπουδών - 1ο Πεδίο
       if values["Επιλογή Κατεύθυνσης"] == "Ανθρωπιστικών Σπουδών - 1ο Πεδίο":
          wb = openpyxl.load_workbook('Αρχεία Εφαρμογής/Moria.xlsx')
          sheet = wb.active
          for i in range((len(pedio1))):
             Y=[]
             for j in range(4,8):
                Y.append(X[j-4]*pedio1.iloc[i,j])
             sheet.append((pedio1.iloc[i,2],pedio1.iloc[i,1],sum(Y)*1000,sum(X)/4,pedio1.iloc[i,11]))
          wb.save('Αρχείο Υπολογισμού Μορίων/'+codename)
          #### Ταξινόμηση σε Φθίνουσα σειρά
          df = pd.read_excel('Αρχείο Υπολογισμού Μορίων/'+codename)
          df = df.sort_values(by="ΜΟΡΙΑ",ascending = False)
          #### Δημιουργία και αποθήκευση ταξινομημένου αρχείου          
          wb1 = openpyxl.load_workbook('Αρχεία Εφαρμογής/Moria.xlsx')
          sheet1 = wb1.active
          for i in range((len(pedio1))):
              sheet1.append((df.iloc[i,0],df.iloc[i,1],df.iloc[i,2],df.iloc[i,3],df.iloc[i,4]))
          wb1.save('Αρχείο Υπολογισμού Μορίων/'+codename)

######## Θετικές και Τεχνολογικές Επιστήμες - 2ο Πεδίο
       elif values["Επιλογή Κατεύθυνσης"] == "Θετικές και Τεχνολογικές Επιστήμες - 2ο Πεδίο":
          wb = openpyxl.load_workbook('Αρχεία Εφαρμογής/Moria.xlsx')
          sheet = wb.active 
          for i in range((len(pedio2))):
             Y=[]
             for j in range(4,8):
                Y.append(X[j-4]*pedio2.iloc[i,j])
             sheet.append((pedio2.iloc[i,2],pedio2.iloc[i,1],sum(Y)*1000,sum(X)/4,pedio2.iloc[i,11]))
          wb.save('Αρχείο Υπολογισμού Μορίων/'+codename)
          #### Ταξινόμηση σε Φθίνουσα σειρά
          df = pd.read_excel('Αρχείο Υπολογισμού Μορίων/'+codename)
          df = df.sort_values(by="ΜΟΡΙΑ",ascending = False)
          #### Δημιουργία και αποθήκευση ταξινομημένου αρχείου          
          wb1 = openpyxl.load_workbook('Αρχεία Εφαρμογής/Moria.xlsx')
          sheet1 = wb1.active
          for i in range((len(pedio2))):
              sheet1.append((df.iloc[i,0],df.iloc[i,1],df.iloc[i,2],df.iloc[i,3],df.iloc[i,4]))
          wb1.save('Αρχείο Υπολογισμού Μορίων/'+codename)
          
######## Επιστήμες Υγείας και Ζωής - 3ο Πεδίο
       elif values["Επιλογή Κατεύθυνσης"] == "Επιστήμες Υγείας και Ζωής - 3ο Πεδίο":
          wb = openpyxl.load_workbook('Αρχεία Εφαρμογής/Moria.xlsx')
          sheet = wb.active 
          for i in range((len(pedio3))):
             Y=[]
             for j in range(4,8):
                Y.append(X[j-4]*pedio3.iloc[i,j])
             sheet.append((pedio3.iloc[i,2],pedio3.iloc[i,1],sum(Y)*1000,sum(X)/4,pedio3.iloc[i,11]))
          wb.save('Αρχείο Υπολογισμού Μορίων/'+codename)
          #### Ταξινόμηση σε Φθίνουσα σειρά
          df = pd.read_excel('Αρχείο Υπολογισμού Μορίων/'+codename)
          df = df.sort_values(by="ΜΟΡΙΑ",ascending = False)
          #### Δημιουργία και αποθήκευση ταξινομημένου αρχείου          
          wb1 = openpyxl.load_workbook('Αρχεία Εφαρμογής/Moria.xlsx')
          sheet1 = wb1.active
          for i in range((len(pedio3))):
              sheet1.append((df.iloc[i,0],df.iloc[i,1],df.iloc[i,2],df.iloc[i,3],df.iloc[i,4]))
          wb1.save('Αρχείο Υπολογισμού Μορίων/'+codename)

######## Επιστήμες Οικονομίας και Πληροφορικής - 4ο Πεδίο          
       elif values["Επιλογή Κατεύθυνσης"] == "Επιστήμες Οικονομίας και Πληροφορικής - 4ο Πεδίο":
          wb = openpyxl.load_workbook('Αρχεία Εφαρμογής/Moria.xlsx')
          sheet = wb.active 
          for i in range((len(pedio4))):
             Y=[]
             for j in range(4,8):
                Y.append(X[j-4]*pedio4.iloc[i,j])
             sheet.append((pedio4.iloc[i,2],pedio4.iloc[i,1],sum(Y)*1000,sum(X)/4,pedio4.iloc[i,11]))
          wb.save('Αρχείο Υπολογισμού Μορίων/'+codename)
          #### Ταξινόμηση σε Φθίνουσα σειρά
          df = pd.read_excel('Αρχείο Υπολογισμού Μορίων/'+codename)
          df = df.sort_values(by="ΜΟΡΙΑ",ascending = False)
          #### Δημιουργία και αποθήκευση ταξινομημένου αρχείου          
          wb1 = openpyxl.load_workbook('Αρχεία Εφαρμογής/Moria.xlsx')
          sheet1 = wb1.active
          for i in range((len(pedio4))):
              sheet1.append((df.iloc[i,0],df.iloc[i,1],df.iloc[i,2],df.iloc[i,3],df.iloc[i,4]))
          wb1.save('Αρχείο Υπολογισμού Μορίων/'+codename)
       #### Άνοιγμα αρχείου excel για τον κάθε μαθητή
       os.startfile(os.getcwd()+'/Αρχείο Υπολογισμού Μορίων/'+codename, 'print')
 
        
window.close()
